"""
ARIMA-Informed LSTM — Web App
==============================
Gradio app for Hugging Face Spaces.
Nav: Home | Model | Instructions | Developers
Model workflow: Upload Data -> Summary Stats & Plots -> ARIMA -> LSTM -> ARIMA-Informed LSTM
"""
import warnings, time, copy, io, os, base64
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error
from statsmodels.tsa.arima.model import ARIMA
import pmdarima as pm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import gradio as gr
from skopt import gp_minimize
from skopt.space import Real, Integer
from skopt.utils import use_named_args

np.random.seed(42)
torch.manual_seed(42)

# ═══════════════════════════════════════════════════════════
# Core model classes (ported from arima_informed_lstm.py)
# ═══════════════════════════════════════════════════════════

class LSTMModel(nn.Module):
    def __init__(self, hidden_size=32, num_layers=1, dropout=0.0):
        super().__init__()
        hidden_size = int(hidden_size)
        num_layers = int(num_layers)
        dropout = float(dropout)
        self.lstm = nn.LSTM(1, hidden_size, num_layers, batch_first=True,
                             dropout=dropout if num_layers > 1 else 0.0)
        self.fc = nn.Linear(hidden_size, 1)

    def forward(self, x):
        out, _ = self.lstm(x)
        return self.fc(out[:, -1, :])


class ARIMAInformedLoss(nn.Module):
    def __init__(self, lambda_arima=0.1):
        super().__init__()
        self.lam = lambda_arima
        self.mse = nn.MSELoss()

    def forward(self, pred, actual, arima_b):
        la = self.mse(pred, actual)
        lb = self.mse(pred, arima_b)
        return (1 - self.lam) * la + self.lam * lb, la, lb


def format_time(s):
    if s < 60:
        return f"{s:.0f}s"
    return f"{int(s // 60)}m {s % 60:.0f}s"

def df_to_html_table(df, max_height=None):
    """Render a small results DataFrame as a plain HTML table — avoids a
    Gradio Dataframe rendering quirk that can silently clip rows in wide tables,
    and avoids the interactive Dataframe grid hanging on large tables."""
    headers = "".join(f"<th style='padding:6px 10px;border:1px solid #ddd;background:#1F4E79;color:#fff;text-align:center;position:sticky;top:0'>{c}</th>"
                       for c in df.columns)
    rows_html = ""
    for i, row in enumerate(df.itertuples(index=False)):
        bg = "#f5f7fa" if i % 2 == 0 else "#ffffff"
        cells = "".join(f"<td style='padding:6px 10px;border:1px solid #ddd;text-align:center'>{v}</td>" for v in row)
        rows_html += f"<tr style='background:{bg}'>{cells}</tr>"
    wrap_style = f"max-height:{max_height}px;overflow-y:auto;" if max_height else ""
    return (f"<div style='overflow-x:auto;{wrap_style}'><table style='border-collapse:collapse;width:100%;font-family:monospace;font-size:0.85rem'>"
            f"<thead><tr>{headers}</tr></thead><tbody>{rows_html}</tbody></table></div>")

def make_seq(s, q):
    X, y = [], []
    for i in range(len(s) - q):
        X.append(s[i:i + q])
        y.append(s[i + q])
    return np.array(X, dtype=np.float32), np.array(y, dtype=np.float32)


def evaluate(model, Xt, yt, scaler):
    model.eval()
    with torch.no_grad():
        ps = model(Xt).squeeze().numpy()
    pr = scaler.inverse_transform(ps.reshape(-1, 1)).flatten()
    ac = scaler.inverse_transform(yt.squeeze().numpy().reshape(-1, 1)).flatten()
    rmse = float(np.sqrt(mean_squared_error(ac, pr)))
    mae = float(mean_absolute_error(ac, pr))
    mask = ac != 0
    mape = float(np.mean(np.abs((ac[mask] - pr[mask]) / ac[mask])) * 100) if mask.any() else 0.0
    return pr, ac, rmse, mae, mape


def count_p(m):
    return sum(p.numel() for p in m.parameters())


def train_lstm(hp, train_loader, Xv, yv, Xt, yt, scaler, arima_tr_t, lam,
                epochs=200, patience=20, batch_size=16):
    model = LSTMModel(hp["hidden"], hp["layers"], hp["dropout"])
    crit = ARIMAInformedLoss(lam)
    optim = torch.optim.Adam(model.parameters(), lr=hp["lr"])
    mse_only = nn.MSELoss()
    best_val = float("inf")
    counter = 0
    best_state = copy.deepcopy(model.state_dict())
    hist = {"total": [], "actual": [], "arima": []}
    bs = min(batch_size, len(train_loader.dataset))

    if lam == 0.0:
        for epoch in range(epochs):
            model.train()
            ep = 0.0
            for bX, by in train_loader:
                optim.zero_grad()
                loss, _, _ = crit(model(bX), by, by)
                loss.backward()
                nn.utils.clip_grad_norm_(model.parameters(), 1.0)
                optim.step()
                ep += loss.item()
            ep /= len(train_loader)
            hist["total"].append(ep); hist["actual"].append(ep); hist["arima"].append(0.0)
            model.eval()
            with torch.no_grad():
                vl = mse_only(model(Xv), yv).item()
            if vl < best_val:
                best_val = vl; counter = 0; best_state = copy.deepcopy(model.state_dict())
            else:
                counter += 1
            if counter >= patience:
                break
    else:
        Xtr = train_loader.dataset.tensors[0]
        ytr = train_loader.dataset.tensors[1]
        ai_dl = DataLoader(TensorDataset(Xtr, ytr, arima_tr_t), batch_size=bs, shuffle=False)
        for epoch in range(epochs):
            model.train()
            et = ea = er = 0.0
            for bX, by, ba in ai_dl:
                optim.zero_grad()
                loss, la, lb = crit(model(bX), by, ba)
                loss.backward()
                nn.utils.clip_grad_norm_(model.parameters(), 1.0)
                optim.step()
                et += loss.item(); ea += la.item(); er += lb.item()
            n = len(ai_dl)
            hist["total"].append(et / n); hist["actual"].append(ea / n); hist["arima"].append(er / n)
            model.eval()
            with torch.no_grad():
                vl = mse_only(model(Xv), yv).item()
            if vl < best_val:
                best_val = vl; counter = 0; best_state = copy.deepcopy(model.state_dict())
            else:
                counter += 1
            if counter >= patience:
                break

    model.load_state_dict(best_state)
    pr, ac, rmse, mae, mape = evaluate(model, Xt, yt, scaler)
    return model, pr, ac, rmse, mae, mape, hist, count_p(model)


def grid_search(param_grid, train_loader, Xv, yv, scaler, tune_epochs=30, progress_fn=None):
    results = []
    combos = [(h, l, lr, d)
              for h in param_grid["hidden_size"]
              for l in param_grid["num_layers"]
              for lr in param_grid["lr"]
              for d in param_grid["dropout"]]
    total = len(combos)
    start_t = time.time()
    for i, (h, l, lr, d) in enumerate(combos):
        if progress_fn is not None:
            elapsed = time.time() - start_t
            avg = elapsed / i if i > 0 else 0.0
            eta = avg * (total - i)
            eta_str = f", ETA {format_time(eta)}" if i > 0 else ""
            progress_fn(i / total, desc=f"Grid search: combo {i + 1}/{total} "
                                        f"(hidden={int(h)}, layers={int(l)}, lr={lr:.4f}, dropout={d:.2f}) "
                                        f"— elapsed {format_time(elapsed)}{eta_str}")
        m = LSTMModel(int(h), int(l), float(d))
        opt = torch.optim.Adam(m.parameters(), lr=float(lr))
        crit = ARIMAInformedLoss(0.0)
        m.train()
        for _ in range(tune_epochs):
            for bX, by in train_loader:
                opt.zero_grad()
                loss, _, _ = crit(m(bX), by, by)
                loss.backward()
                opt.step()
        _, _, rv, _, _ = evaluate(m, Xv, yv, scaler)
        results.append(dict(hidden=int(h), layers=int(l), lr=float(lr), dropout=float(d), rmse=rv))
    if progress_fn is not None:
        progress_fn(1.0, desc=f"Grid search: {total}/{total} combos complete "
                              f"in {format_time(time.time() - start_t)}")
    return sorted(results, key=lambda x: x["rmse"])[0], results


def bayes_search(train_loader, Xv, yv, scaler, hidden_range, layers_range, lr_range, dropout_range,
                  n_calls=15, tune_epochs=30, random_state=42, progress_fn=None):
    """True Bayesian optimization: fits a Gaussian Process surrogate over the
    hyperparameter space and uses expected-improvement to choose each next
    evaluation, via scikit-optimize's gp_minimize."""

    # Any hyperparameter whose min == max is held fixed and excluded from the
    # search space (skopt requires a genuine range for each dimension).
    specs, fixed = [], {}

    def add_dim(name, low, high, is_int):
        low = float(low); high = float(high)
        if low > high:
            low, high = high, low
        if is_int:
            low, high = int(round(low)), int(round(high))
        if low == high:
            fixed[name] = low
        else:
            specs.append((name, low, high, is_int))

    add_dim("hidden", hidden_range[0], hidden_range[1], is_int=True)
    add_dim("layers", layers_range[0], layers_range[1], is_int=True)
    add_dim("lr", lr_range[0], lr_range[1], is_int=False)
    add_dim("dropout", dropout_range[0], dropout_range[1], is_int=False)

    def _train_and_eval(hidden, layers, lr, dropout):
        m = LSTMModel(int(round(hidden)), int(round(layers)), float(dropout))
        opt = torch.optim.Adam(m.parameters(), lr=float(lr))
        crit = ARIMAInformedLoss(0.0)
        m.train()
        for _ in range(tune_epochs):
            for bX, by in train_loader:
                opt.zero_grad()
                loss, _, _ = crit(m(bX), by, by)
                loss.backward()
                opt.step()
        _, _, rv, _, _ = evaluate(m, Xv, yv, scaler)
        return rv

    if not specs:
        # Everything is fixed — nothing to optimize, just one evaluation.
        if progress_fn is not None:
            progress_fn(0.0, desc="Evaluating fixed hyperparameters...")
        rv = _train_and_eval(fixed["hidden"], fixed["layers"], fixed["lr"], fixed["dropout"])
        if progress_fn is not None:
            progress_fn(1.0, desc="Done")
        best = dict(hidden=int(fixed["hidden"]), layers=int(fixed["layers"]),
                    lr=float(fixed["lr"]), dropout=float(fixed["dropout"]), rmse=rv)
        return best, [best]

    dimensions = [Integer(low, high, name=name) if is_int else Real(low, high, name=name)
                  for name, low, high, is_int in specs]

    @use_named_args(dimensions)
    def objective(**params):
        p = {**fixed, **params}
        return _train_and_eval(p["hidden"], p["layers"], p["lr"], p["dropout"])

    n_calls = max(int(n_calls), len(dimensions) + 2)
    n_initial = max(2, min(5, n_calls // 2))  # random warm-start points before GP takes over
    bayes_start_t = time.time()

    def _gp_progress_callback(res):
        if progress_fn is not None:
            done = len(res.x_iters)
            elapsed = time.time() - bayes_start_t
            avg = elapsed / done if done > 0 else 0.0
            eta = avg * (n_calls - done)
            eta_str = f", ETA {format_time(eta)}" if done < n_calls else ""
            progress_fn(done / n_calls, desc=f"Bayesian optimization: evaluation {done}/{n_calls} "
                                              f"(best RMSE so far: {min(res.func_vals):.4f}) "
                                              f"— elapsed {format_time(elapsed)}{eta_str}")

    result = gp_minimize(objective, dimensions, n_calls=n_calls,
                          n_initial_points=n_initial, random_state=random_state, verbose=False,
                          callback=_gp_progress_callback if progress_fn is not None else None)

    def _to_hp(x_vals, rmse):
        params = {name: v for (name, *_), v in zip(specs, x_vals)}
        params = {**fixed, **params}
        return dict(hidden=int(round(params["hidden"])), layers=int(round(params["layers"])),
                    lr=float(params["lr"]), dropout=float(params["dropout"]), rmse=float(rmse))

    best = _to_hp(result.x, result.fun)
    all_results = [_to_hp(x, y) for x, y in zip(result.x_iters, result.func_vals)]
    return best, all_results


def build_grid_values(mn, mx, step, is_int):
    """Build an inclusive list of values from a min/max/step spec (used for Grid Search)."""
    mn, mx, step = float(mn), float(mx), float(step)
    if mx < mn:
        mn, mx = mx, mn
    if step <= 0:
        step = 1.0 if is_int else 0.05
    vals = np.arange(mn, mx + step / 2, step)
    if is_int:
        vals = np.unique(np.round(vals).astype(int))
    else:
        vals = np.unique(np.round(vals, 6))
    return vals.tolist()


# ═══════════════════════════════════════════════════════════
# App-level state helpers
# ═══════════════════════════════════════════════════════════

def load_file(file):
    if file is None:
        return None, gr.update(choices=[]), gr.update(choices=[]), "Upload a CSV or Excel file to begin."
    path = file.name if hasattr(file, "name") else file
    if str(path).lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)
    cols = list(df.columns)
    msg = f"Loaded {len(df)} rows, {len(cols)} columns: {', '.join(cols)}"
    return df, gr.update(choices=cols, value=cols[0] if cols else None), \
           gr.update(choices=cols, value=cols[1] if len(cols) > 1 else None), msg


def summary_stats(df, time_col, value_col, stats_selected, show_line, show_box,
                   plot_width=8, plot_height=3.5, font_size=9):
    if df is None or value_col is None:
        return "Please upload data and select a study variable first.", None, None
    s = pd.to_numeric(df[value_col], errors="coerce").dropna()
    stat_map = {
        "Mean": s.mean(), "Median": s.median(), "Standard Deviation": s.std(),
        "Variance": s.var(), "Minimum": s.min(), "Maximum": s.max(),
        "Skewness": s.skew(), "Kurtosis": s.kurt(),
    }
    lines = [f"**{k}**: {stat_map[k]:.3f}" for k in stats_selected if k in stat_map]
    text = "\n\n".join(lines) if lines else "Select at least one statistic to display."

    line_fig = None
    if show_line:
        import matplotlib.dates as mdates
        line_fig, ax = plt.subplots(figsize=(plot_width, plot_height))
        raw_x = df[time_col][:len(s)] if time_col in df.columns else pd.Series(range(len(s)))
        x_axis = None
        if not pd.api.types.is_numeric_dtype(raw_x):
            parsed = pd.to_datetime(raw_x, errors="coerce", dayfirst=True)
            if parsed.notna().all() and parsed.dt.year.between(1678, 2262).all():
                x_axis = parsed
        if x_axis is not None:
            order = np.argsort(x_axis.values)
            ax.plot(x_axis.values[order], s.values[order], color="#1F4E79", linewidth=1.5)
            locator = mdates.AutoDateLocator(minticks=5, maxticks=12)
            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
            plt.setp(ax.get_xticklabels(), rotation=90, ha="center")
        else:
            x_pos = np.arange(len(s))
            ax.plot(x_pos, s.values, color="#1F4E79", linewidth=1.5)
            step = max(1, len(s) // 12)
            tick_idx = x_pos[::step]
            ax.set_xticks(tick_idx)
            ax.set_xticklabels([str(raw_x.iloc[i]) for i in tick_idx], rotation=90, ha="center")
        ax.tick_params(axis='both', labelsize=font_size)
        ax.set_title(f"{value_col} over time", fontsize=font_size + 2)
        ax.set_xlabel(time_col, fontsize=font_size); ax.set_ylabel(value_col, fontsize=font_size)
        line_fig.tight_layout()

    box_fig = None
    if show_box:
        box_fig, ax = plt.subplots(figsize=(plot_width, plot_height))
        ax.boxplot(s.values, patch_artist=True,
                   boxprops=dict(facecolor="#D9E1F2"))
        ax.set_title(f"{value_col} — Boxplot", fontsize=font_size + 2)
        ax.tick_params(axis='both', labelsize=font_size)
        box_fig.tight_layout()

    return text, line_fig, box_fig


def run_arima(df, value_col, use_auto, p, d, q, train_ratio):
    if df is None or value_col is None:
        return "Upload data first.", None
    data = pd.to_numeric(df[value_col], errors="coerce").dropna().values.astype(float)
    N = len(data)
    split = int(N * train_ratio)
    if N - split < 5:
        split = N - 5
    train_data, test_data = data[:split], data[split:]

    if use_auto:
        auto = pm.auto_arima(train_data, start_p=0, max_p=5, start_q=0, max_q=5,
                              d=None, seasonal=False, stepwise=True,
                              suppress_warnings=True, error_action="ignore")
        order = auto.order
    else:
        order = (int(p), int(d), int(q))

    model = ARIMA(train_data, order=order).fit()
    forecast = np.array(model.forecast(steps=len(test_data)))
    rmse = float(np.sqrt(mean_squared_error(test_data, forecast)))
    mae = float(mean_absolute_error(test_data, forecast))
    mask = test_data != 0
    mape = float(np.mean(np.abs((test_data[mask] - forecast[mask]) / test_data[mask])) * 100) if mask.any() else 0.0

    summary = (f"**Order (p,d,q)**: {order}\n\n"
               f"**Test RMSE**: {rmse:.3f}\n\n**Test MAE**: {mae:.3f}\n\n**Test MAPE**: {mape:.3f}%")
    return summary, dict(order=order, rmse=rmse, mae=mae, mape=mape,
                          preds=forecast.tolist(), actuals=test_data.tolist())


def run_lstm_tuning(df, value_col, seq_len, train_ratio, tune_mode,
                     # Manual
                     hidden, layers, lr, dropout,
                     # Grid Search: min, max, step for each hyperparameter
                     g_hidden_min, g_hidden_max, g_hidden_step,
                     g_layers_min, g_layers_max, g_layers_step,
                     g_lr_min, g_lr_max, g_lr_step,
                     g_dropout_min, g_dropout_max, g_dropout_step,
                     # Bayesian Search: min, max for each hyperparameter + iterations
                     b_hidden_min, b_hidden_max,
                     b_layers_min, b_layers_max,
                     b_lr_min, b_lr_max,
                     b_dropout_min, b_dropout_max,
                     n_calls,
                     progress=gr.Progress()):
    if df is None or value_col is None:
        return "Upload data first.", None
    np.random.seed(42); torch.manual_seed(42)  # reproducible across repeated clicks
    progress(0, desc="Preparing data...")
    data = pd.to_numeric(df[value_col], errors="coerce").dropna().values.astype(float)
    sc = MinMaxScaler()
    scaled = sc.fit_transform(data.reshape(-1, 1)).flatten()
    X_all, y_all = make_seq(scaled, int(seq_len))
    N = len(data)
    split = int(N * train_ratio)
    if N - split < 5:
        split = N - 5
    seq_split = split - int(seq_len)
    X_train, y_train = X_all[:seq_split], y_all[:seq_split]
    val_split = int(len(X_train) * 0.8)
    X_tr, X_val = X_train[:val_split], X_train[val_split:]
    y_tr, y_val = y_train[:val_split], y_train[val_split:]

    Xtr_t = torch.FloatTensor(X_tr).unsqueeze(-1); ytr_t = torch.FloatTensor(y_tr).unsqueeze(-1)
    Xv_t = torch.FloatTensor(X_val).unsqueeze(-1); yv_t = torch.FloatTensor(y_val).unsqueeze(-1)
    bs = min(16, len(X_tr))
    train_loader = DataLoader(TensorDataset(Xtr_t, ytr_t), batch_size=bs, shuffle=True)

    if tune_mode == "Manual":
        progress(0.5, desc="Applying manual hyperparameters...")
        best_hp = dict(hidden=int(hidden), layers=int(layers), lr=float(lr), dropout=float(dropout))
        progress(1.0, desc="Done")
        msg = (f"**Manual hyperparameters set** — hidden={best_hp['hidden']}, "
               f"layers={best_hp['layers']}, lr={best_hp['lr']}, dropout={best_hp['dropout']}")

    elif tune_mode == "Grid Search":
        param_grid = {
            "hidden_size": build_grid_values(g_hidden_min, g_hidden_max, g_hidden_step, is_int=True),
            "num_layers": build_grid_values(g_layers_min, g_layers_max, g_layers_step, is_int=True),
            "lr": build_grid_values(g_lr_min, g_lr_max, g_lr_step, is_int=False),
            "dropout": build_grid_values(g_dropout_min, g_dropout_max, g_dropout_step, is_int=False),
        }
        n_combos = (len(param_grid["hidden_size"]) * len(param_grid["num_layers"]) *
                    len(param_grid["lr"]) * len(param_grid["dropout"]))
        best_hp, _ = grid_search(param_grid, train_loader, Xv_t, yv_t, sc, progress_fn=progress)
        msg = (f"**Grid search complete** ({n_combos} combinations tried) — "
               f"best: hidden={best_hp['hidden']}, layers={best_hp['layers']}, "
               f"lr={best_hp['lr']:.4f}, dropout={best_hp['dropout']:.2f} (val RMSE={best_hp['rmse']:.4f})")

    else:  # Bayesian Search
        best_hp, _ = bayes_search(
            train_loader, Xv_t, yv_t, sc,
            hidden_range=(int(b_hidden_min), int(b_hidden_max)),
            layers_range=(int(b_layers_min), int(b_layers_max)),
            lr_range=(float(b_lr_min), float(b_lr_max)),
            dropout_range=(float(b_dropout_min), float(b_dropout_max)),
            n_calls=int(n_calls), progress_fn=progress)
        msg = (f"**Bayesian optimization complete** (Gaussian Process, {int(n_calls)} evaluations) — "
               f"best: hidden={best_hp['hidden']}, layers={best_hp['layers']}, "
               f"lr={best_hp['lr']:.4f}, dropout={best_hp['dropout']:.2f} (val RMSE={best_hp['rmse']:.4f})")

    return msg, best_hp


def run_full_pipeline(df, value_col, seq_len, train_ratio, best_hp, lambdas_str,
                       epochs, patience):
    if df is None or value_col is None or best_hp is None:
        return "Complete the Data, ARIMA, and LSTM steps first.", None, None, None, None, None, None

    lambdas = [0.0] + sorted(set(float(x.strip()) for x in lambdas_str.split(",") if x.strip() != ""))
    data = pd.to_numeric(df[value_col], errors="coerce").dropna().values.astype(float)
    N = len(data)
    split = int(N * train_ratio)
    if N - split < 5:
        split = N - 5
    train_data, test_data = data[:split], data[split:]

    # ARIMA fitted/forecast for the informed loss term
    order = pm.auto_arima(train_data, seasonal=False, stepwise=True,
                           suppress_warnings=True, error_action="ignore").order
    am = ARIMA(train_data, order=order).fit()
    arima_fitted = np.array(am.fittedvalues)
    arima_test = np.array(am.forecast(steps=len(test_data)))

    sc = MinMaxScaler()
    scaled = sc.fit_transform(data.reshape(-1, 1)).flatten()
    seq_len = int(seq_len)
    X_all, y_all = make_seq(scaled, seq_len)
    arima_test_sc = sc.transform(arima_test.reshape(-1, 1)).flatten().astype(np.float32)
    seq_split = split - seq_len
    X_train, y_train = X_all[:seq_split], y_all[:seq_split]
    X_test, y_test = X_all[seq_split:], y_all[seq_split:]
    val_split = int(len(X_train) * 0.8)
    X_tr, X_val = X_train[:val_split], X_train[val_split:]
    y_tr, y_val = y_train[:val_split], y_train[val_split:]

    Xtr_t = torch.FloatTensor(X_tr).unsqueeze(-1); ytr_t = torch.FloatTensor(y_tr).unsqueeze(-1)
    Xv_t = torch.FloatTensor(X_val).unsqueeze(-1); yv_t = torch.FloatTensor(y_val).unsqueeze(-1)
    Xte_t = torch.FloatTensor(X_test).unsqueeze(-1); yte_t = torch.FloatTensor(y_test).unsqueeze(-1)
    bs = min(16, len(X_tr))
    train_loader = DataLoader(TensorDataset(Xtr_t, ytr_t), batch_size=bs, shuffle=True)

    af = arima_fitted.copy()
    req = seq_len + len(y_train)
    if len(af) < req:
        af = np.concatenate([af, np.full(req - len(af), af[-1])])
    aln = af[seq_len:seq_len + len(y_train)]
    asc = sc.transform(aln.reshape(-1, 1)).flatten().astype(np.float32)
    arima_tr_t = torch.FloatTensor(asc[:val_split]).unsqueeze(-1)

    Xtr_full_t = torch.FloatTensor(X_train).unsqueeze(-1)
    ytr_full_t = torch.FloatTensor(y_train).unsqueeze(-1)

    rows = []
    all_preds = {}
    for lam in lambdas:
        np.random.seed(42); torch.manual_seed(42)
        t0 = time.time()
        model, pr, ac, rmse, mae, mape, hist, npar = train_lstm(
            best_hp, train_loader, Xv_t, yv_t, Xte_t, yte_t, sc, arima_tr_t, lam,
            epochs=int(epochs), patience=int(patience))
        elapsed = time.time() - t0
        # Evaluate the same trained model on the full training set (train+val)
        tr_pr, tr_ac, tr_rmse, tr_mae, tr_mape = evaluate(model, Xtr_full_t, ytr_full_t, sc)
        label = "Standard (λ=0)" if lam == 0.0 else f"λ={lam}"
        rows.append([label, round(tr_rmse, 3), round(tr_mae, 3), round(tr_mape, 3),
                     round(rmse, 3), round(mae, 3), round(mape, 3), npar, round(elapsed, 2)])
        all_preds[label] = dict(train=tr_pr, test=pr)

    results_df = pd.DataFrame(rows, columns=["Model", "Train RMSE", "Train MAE", "Train MAPE (%)",
                                              "Test RMSE", "Test MAE", "Test MAPE (%)", "Params", "Time (s)"])

    # Actual values (same regardless of lambda) and ARIMA train-fit/test-forecast,
    # all in original scale, concatenated into one continuous train+test series.
    train_actual_inv = sc.inverse_transform(y_train.reshape(-1, 1)).flatten()
    test_actual_inv = sc.inverse_transform(y_test.reshape(-1, 1)).flatten()
    full_actual = np.concatenate([train_actual_inv, test_actual_inv])
    full_arima = np.concatenate([aln, arima_test[:len(test_actual_inv)]])
    split_x = len(train_actual_inv)

    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(full_actual, label="Actual", color="black", linewidth=2)
    ax.plot(full_arima, label="ARIMA", linestyle=":", color="gray")
    colors = plt.cm.viridis(np.linspace(0, 1, len(all_preds)))
    for (label, preds), c in zip(all_preds.items(), colors):
        full_pred = np.concatenate([preds["train"], preds["test"]])
        ax.plot(full_pred, label=label, color=c, alpha=0.8)
    ax.axvline(x=split_x - 0.5, color="red", linestyle="--", linewidth=1.2, label="Train / Test split")
    ax.legend(fontsize=7, ncol=2)
    ax.set_title("ARIMA-Informed LSTM — Train & Test Predictions")
    ax.set_xlabel("Time Step"); ax.set_ylabel(value_col)
    fig.tight_layout()

    # Per-timestep Actual vs Predicted tables (one column per lambda), separately
    # for train and test — this is the raw data behind the plot above.
    train_table = {"Time Step": np.arange(1, len(train_actual_inv) + 1),
                    "Actual": np.round(train_actual_inv, 3),
                    "ARIMA": np.round(aln, 3)}
    for label, preds in all_preds.items():
        train_table[label] = np.round(preds["train"], 3)
    train_pred_df = pd.DataFrame(train_table)

    test_table = {"Time Step": np.arange(1, len(test_actual_inv) + 1),
                   "Actual": np.round(test_actual_inv, 3),
                   "ARIMA": np.round(arima_test[:len(test_actual_inv)], 3)}
    for label, preds in all_preds.items():
        test_table[label] = np.round(preds["test"], 3)
    test_pred_df = pd.DataFrame(test_table)

    # Configuration / hyperparameters summary — everything that went into this run.
    config_rows = [
        ["Sequence Length (lag window)", int(seq_len)],
        ["Train / Test Split Ratio", round(float(train_ratio), 2)],
        ["Train Set Size", int(len(train_actual_inv))],
        ["Test Set Size", int(len(test_actual_inv))],
        ["ARIMA Order (p, d, q)", str(order)],
        ["LSTM Hidden Size", int(best_hp["hidden"])],
        ["LSTM Num Layers", int(best_hp["layers"])],
        ["LSTM Learning Rate", float(best_hp["lr"])],
        ["LSTM Dropout", float(best_hp["dropout"])],
        ["λ Values Tested", ", ".join(f"{lam:g}" for lam in lambdas)],
        ["Max Epochs", int(epochs)],
        ["Early-Stopping Patience", int(patience)],
    ]
    config_df = pd.DataFrame(config_rows, columns=["Parameter", "Value"])

    # Excel export — save under a local outputs/ folder (cross-platform, always
    # inside an allowed Gradio path) rather than a hardcoded /tmp path.
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "ARIMA_Informed_LSTM_Results.xlsx")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Summary")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for ci, h in enumerate(results_df.columns, 1):
        c = ws.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
        c.alignment = Alignment(horizontal="center"); c.border = border
    for ri, row in enumerate(results_df.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(ri, ci, v); c.border = border; c.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEFGHI", [18, 12, 12, 14, 12, 12, 14, 10, 10]):
        ws.column_dimensions[col].width = w

    def _write_pred_sheet(ws2, table_df):
        for ci, h in enumerate(table_df.columns, 1):
            c = ws2.cell(1, ci, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
            c.alignment = Alignment(horizontal="center"); c.border = border
        for ri, row in enumerate(table_df.itertuples(index=False), 2):
            for ci, v in enumerate(row, 1):
                c = ws2.cell(ri, ci, v); c.border = border; c.alignment = Alignment(horizontal="center")
        for ci in range(1, len(table_df.columns) + 1):
            ws2.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = 13

    ws_config = wb.create_sheet("Config")
    _write_pred_sheet(ws_config, config_df)
    ws_config.column_dimensions["A"].width = 30
    ws_config.column_dimensions["B"].width = 24

    ws_train = wb.create_sheet("Train_Predictions")
    _write_pred_sheet(ws_train, train_pred_df)
    ws_test = wb.create_sheet("Test_Predictions")
    _write_pred_sheet(ws_test, test_pred_df)

    wb.save(out_path)

    return ("Training complete for all λ values.", df_to_html_table(results_df), fig, out_path,
             df_to_html_table(train_pred_df, max_height=400), df_to_html_table(test_pred_df, max_height=400),
             config_df)


# ═══════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════

STAT_CHOICES = ["Mean", "Median", "Standard Deviation", "Variance",
                 "Minimum", "Maximum", "Skewness", "Kurtosis"]

# ─── Developers page data ───
ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")

DEVELOPERS = [
    dict(name="Ranjit Kumar Paul", role="National Fellow",
         affiliation=["ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="ranjitstat@gmail.com",
         scholar="https://scholar.google.com/citations?hl=en&user=wBWuZJgAAAAJ&view_op=list_works&sortby=pubdate",
         photo="ranjit_kumar_paul.png"),
    dict(name="Md Yeasin", role="Scientist",
         affiliation=["Division of Statistical Ecology and Environmental Statistics",
                      "ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="yeasin.iasri@gmail.com",
         scholar="https://scholar.google.com/citations?user=xejMKD0AAAAJ",
         photo="md_yeasin.png"),
    dict(name="Pushkar Bora", role="Research Fellow",
         affiliation=["Discipline of Agricultural Statistics",
                      "ICAR-Indian Agricultural Statistics Research Institute",
                      "Library Avenue, Pusa, New Delhi, Delhi – 110012, India"],
         email="borapushkar1999@gmail.com",
         scholar="https://scholar.google.com/citations?user=xVIXGlwAAAAJ&hl=en",
         photo="pushkar_bora.jpg"),
]
COUNTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "visitor_count.txt")

def get_and_increment_visitor_count():
    try:
        with open(COUNTER_FILE, "r") as f:
            count = int(f.read().strip() or "0")
    except (FileNotFoundError, ValueError):
        count = 0
    count += 1
    try:
        with open(COUNTER_FILE, "w") as f:
            f.write(str(count))
    except OSError:
        pass
    return count

def _img_to_data_uri(filename):
    path = os.path.join(ASSETS_DIR, filename)
    ext = os.path.splitext(filename)[1].lower()
    mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
    try:
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        return f"data:{mime};base64,{b64}"
    except FileNotFoundError:
        return ""


def build_developers_html():
    cards = []
    for dev in DEVELOPERS:
        img_src = _img_to_data_uri(dev["photo"])
        affil_html = "".join(f"<div class='dev-affil'>{line}</div>" for line in dev["affiliation"])
        cards.append(f"""
        <div class="dev-card">
            <div class="dev-photo-wrap"><img class="dev-photo" src="{img_src}" alt="{dev['name']}"/></div>
            <div class="dev-info">
                <div class="dev-name">{dev['name']}</div>
                <div class="dev-role">{dev['role']}</div>
                {affil_html}
                <div class="dev-links">
                    <a class="dev-btn dev-btn-email" href="mailto:{dev['email']}">✉&nbsp;Email</a>
                    <a class="dev-btn dev-btn-scholar" href="{dev['scholar']}" target="_blank" rel="noopener">🎓&nbsp;Google Scholar</a>
                </div>
            </div>
        </div>""")

    style = """
    <style>
    .dev-wrap { max-width: 980px; margin: 4px auto 20px auto; font-family: inherit; }
    .dev-wrap .dev-intro {
        color: #374151 !important; font-size: 0.97rem; line-height: 1.5;
        margin-bottom: 18px;
    }
    .dev-wrap .dev-grid {
        display: flex; flex-direction: row; flex-wrap: wrap;
        gap: 18px; justify-content: center;
    }
    .dev-wrap .dev-card {
        display: flex; flex-direction: column; align-items: center; text-align: center;
        gap: 12px; flex: 1 1 280px; max-width: 300px;
        background: linear-gradient(145deg, #0b1120, #131c33) !important;
        border: 1px solid #263352 !important; border-radius: 16px !important;
        padding: 24px 20px !important;
        box-shadow: 0 4px 18px rgba(0,0,0,0.28) !important;
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .dev-wrap .dev-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 26px rgba(0,0,0,0.38) !important;
    }
    .dev-wrap .dev-photo-wrap {
        flex-shrink: 0; padding: 2px;
        border-radius: 14px !important;
        background: linear-gradient(135deg, #22d3ee, #6d28d9) !important;
    }
    .dev-wrap .dev-photo {
        display: block; width: 96px !important; height: 96px !important;
        object-fit: cover !important; border-radius: 12px !important;
        border: 2px solid #0b1120 !important;
    }
    .dev-wrap .dev-info { display: flex; flex-direction: column; align-items: center; gap: 2px; min-width: 0; }
    .dev-wrap .dev-name { color: #f8fafc !important; font-size: 1.12rem !important; font-weight: 700 !important; line-height: 1.25; }
    .dev-wrap .dev-role { color: #22d3ee !important; font-weight: 600 !important; font-size: 0.88rem !important; margin-bottom: 3px; }
    .dev-wrap .dev-affil { color: #94a3b8 !important; font-size: 0.8rem !important; line-height: 1.3; }
    .dev-wrap .dev-links { display: flex; gap: 8px; margin-top: 10px; flex-wrap: wrap; justify-content: center; }
    .dev-wrap .dev-btn {
        display: inline-flex !important; align-items: center; gap: 5px;
        padding: 5px 13px !important; border-radius: 999px !important; font-size: 0.78rem !important;
        font-weight: 600 !important; text-decoration: none !important;
        color: #ffffff !important; border: none !important;
    }
    .dev-wrap .dev-btn-email { background: #0e7490 !important; }
    .dev-wrap .dev-btn-email:hover { background: #06b6d4 !important; }
    .dev-wrap .dev-btn-scholar { background: #6d28d9 !important; }
    .dev-wrap .dev-btn-scholar:hover { background: #8b5cf6 !important; }
    </style>
    """
    return style + f'<div class="dev-wrap"><div class="dev-grid">' + "".join(cards) + "</div></div>"



FONT_HEAD = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
"""

CUSTOM_THEME = gr.themes.Base(
    primary_hue=gr.themes.colors.orange,
    secondary_hue=gr.themes.colors.violet,
    neutral_hue=gr.themes.colors.slate,
    font=[gr.themes.GoogleFont("Inter"), "sans-serif"],
    font_mono=[gr.themes.GoogleFont("JetBrains Mono"), "monospace"],
).set(
    body_background_fill="#faf6ee",
    body_background_fill_dark="#faf6ee",
    background_fill_primary="#ffffff",
    background_fill_primary_dark="#ffffff",
    background_fill_secondary="#fdf7ea",
    background_fill_secondary_dark="#fdf7ea",
    border_color_primary="#ecdfc4",
    border_color_primary_dark="#ecdfc4",
    block_background_fill="#ffffff",
    block_background_fill_dark="#ffffff",
    block_border_color="#ecdfc4",
    block_border_color_dark="#ecdfc4",
    block_label_background_fill="#fdf1de",
    block_label_background_fill_dark="#fdf1de",
    block_title_text_color="#b5591a",
    block_title_text_color_dark="#b5591a",
    body_text_color="#2a2418",
    body_text_color_dark="#2a2418",
    body_text_color_subdued="#8a7f68",
    body_text_color_subdued_dark="#8a7f68",
    input_background_fill="#fffdf8",
    input_background_fill_dark="#fffdf8",
    input_border_color="#ecdfc4",
    input_border_color_dark="#ecdfc4",
    button_primary_background_fill="linear-gradient(135deg, #ff9d4d, #e2725b)",
    button_primary_background_fill_dark="linear-gradient(135deg, #ff9d4d, #e2725b)",
    button_primary_text_color="#ffffff",
    button_primary_text_color_dark="#ffffff",
    button_secondary_background_fill="#ffffff",
    button_secondary_background_fill_dark="#ffffff",
    button_secondary_text_color="#2a2418",
    button_secondary_text_color_dark="#2a2418",
    button_secondary_border_color="#ecdfc4",
    button_secondary_border_color_dark="#ecdfc4",
    shadow_drop="0 10px 30px rgba(180,120,40,.14)",
)

CUSTOM_CSS = """
:root {
  --coral: #e2725b; --amber: #f5a742; --teal: #2bb8a8; --violet: #8b5cf6; --ink: #2a2418;
}
@keyframes drift {
  0%   { transform: translate(0,0) scale(1); }
  50%  { transform: translate(3%,-4%) scale(1.06); }
  100% { transform: translate(0,0) scale(1); }
}
.gradio-container {
  font-family: 'Inter', sans-serif !important;
  position: relative;
  background: #faf6ee !important;
  overflow-x: hidden;
}
.gradio-container::before {
  content: ''; position: fixed; inset: -10%; z-index: 0; pointer-events: none;
  background:
    radial-gradient(38% 30% at 12% 8%,  rgba(245,167,66,.35), transparent 70%),
    radial-gradient(35% 28% at 88% 6%,  rgba(139,92,246,.28), transparent 70%),
    radial-gradient(40% 32% at 90% 85%, rgba(43,184,168,.30), transparent 70%),
    radial-gradient(35% 30% at 8% 90%,  rgba(226,114,91,.28), transparent 70%);
  filter: blur(40px);
  animation: drift 22s ease-in-out infinite;
}
.gradio-container > * { position: relative; z-index: 1; }
h1, h2, h3, .prose h1, .prose h2, .prose h3 {
  font-family: 'Fraunces', serif !important; font-weight: 600 !important; letter-spacing: -.01em;
  color: var(--ink) !important;
}
#nav-title {
  margin: -16px -16px 0 -16px !important;
  padding: 0 !important;
  line-height: 0;
}
#nav-title img {
  display: block !important;
  width: 100% !important;
  height: auto !important;
  max-height: none !important;
}
.tabs > .tab-nav {
  border: none !important; gap: 4px; background: #ffffff; border-radius: 999px;
  padding: 4px !important; border: 1px solid #ecdfc4 !important; width: fit-content;
  box-shadow: 0 4px 16px rgba(180,120,40,.10);
}
.tabs > .tab-nav > button {
  border: none !important; border-radius: 999px !important; font-weight: 600 !important;
  color: #8a7f68 !important; padding: 9px 20px !important; margin: 0 !important;
  transition: color .15s ease;
}
.tabs > .tab-nav > button.selected {
  background: linear-gradient(135deg, var(--amber), var(--coral)) !important;
  color: #ffffff !important;
}
button.primary {
  border: none !important; font-weight: 700 !important; border-radius: 999px !important;
  box-shadow: 0 6px 18px rgba(226,114,91,.28);
  transition: transform .15s ease, box-shadow .15s ease;
}
button.primary:hover { transform: translateY(-2px); box-shadow: 0 10px 24px rgba(226,114,91,.36); }
button.secondary {
  background: #ffffff !important;
  border: 2px solid var(--coral) !important;
  color: var(--coral) !important;
  border-radius: 999px !important; font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(226,114,91,.15);
  transition: transform .15s ease, background .15s ease, color .15s ease;
}
button.secondary:hover { transform: translateY(-2px); background: var(--coral) !important; color: #ffffff !important; }
.block {
  border-radius: 18px !important;
  transition: box-shadow .2s ease, transform .2s ease;
}
.gr-box, .form { border-radius: 18px !important; }
input[type="number"], input[type="text"], textarea, select {
  background: #ffffff !important;
  border: 1.5px solid #d9c9a3 !important;
  box-shadow: inset 0 1px 3px rgba(180,120,40,.06) !important;
}
input[type="number"]:focus, input[type="text"]:focus, textarea:focus, select:focus {
  border-color: var(--coral) !important;
  box-shadow: 0 0 0 3px rgba(226,114,91,.15) !important;
}
table { font-family: 'JetBrains Mono', monospace !important; font-size: .82rem !important; }
thead th {
  background: linear-gradient(135deg, #fdf1de, #fbe6d3) !important; color: #b5591a !important;
}
.prose a { color: var(--coral) !important; font-weight: 600; }
footer { opacity: .55; }
/* ── Instructions page — card layout ── */
.instr-section { margin-bottom: 34px; }
.instr-section h3 {
  font-family: 'Fraunces', serif; font-size: 1.3rem; color: var(--ink);
  display: flex; align-items: center; gap: 10px; margin-bottom: 18px;
}
.instr-section h3::before {
  content: ''; width: 26px; height: 2px; background: linear-gradient(90deg, var(--coral), var(--amber));
}
.step-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 14px; }
.step-card {
  background: #ffffff; border: 1px solid #ecdfc4; border-radius: 16px; padding: 18px 20px;
  box-shadow: 0 4px 16px rgba(180,120,40,.08);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.step-card:hover { transform: translateY(-3px); box-shadow: 0 10px 26px rgba(226,114,91,.18); border-color: var(--coral); }
.step-card .step-num {
  display: inline-flex; align-items: center; justify-content: center;
  width: 30px; height: 30px; border-radius: 999px; margin-bottom: 10px;
  background: linear-gradient(135deg, var(--amber), var(--coral)); color: #fff;
  font-family: 'JetBrains Mono', monospace; font-weight: 700; font-size: .85rem;
}
.step-card .step-title { font-family: 'Fraunces', serif; font-weight: 600; font-size: 1rem; color: var(--ink); margin-bottom: 6px; }
.step-card .step-body { font-size: .87rem; color: #6b5c46; line-height: 1.5; }
.note-box {
  background: #fdf1de; border: 1px dashed #f0d9a8; border-radius: 14px; padding: 16px 20px;
  font-style: italic; color: #8a7f68; font-size: .88rem;
}
.methodology-box {
  background: #ffffff; border: 1px solid #ecdfc4; border-radius: 18px;
  padding: 28px 32px; box-shadow: 0 4px 18px rgba(180,120,40,.08);
  position: relative;
}
.methodology-box::before {
  content: ''; position: absolute; top: 0; left: 0; width: 5px; height: 100%;
  background: linear-gradient(180deg, var(--coral), var(--amber));
  border-radius: 18px 0 0 18px;
}
.methodology-box p {
  margin: 0 0 14px; padding-left: 14px; color: #4a4232; font-size: .98rem;
  line-height: 1.9; text-align: justify;
}
.methodology-box p:last-child { margin-bottom: 0; }
.formula-block {
  background: #fdf1de; border: 1px solid #f0d9a8; border-radius: 12px;
  padding: 18px 20px; margin: 4px 14px 16px; text-align: center;
  font-family: 'JetBrains Mono', monospace; font-size: 1.05rem; color: #b5591a;
  font-style: italic; overflow-x: auto;
}
.formula-block--small { font-size: .95rem; }
.frac { display: inline-flex; flex-direction: column; vertical-align: middle; margin: 0 2px; font-size: .8em; line-height: 1; }
.frac-num { border-bottom: 1.4px solid #b5591a; padding: 0 3px 2px; }
.frac-den { padding: 2px 3px 0; }
.sum { font-size: 1.3em; margin: 0 2px; }

/* hero (Home tab) */
.hero-banner { text-align: center; padding: 10px 20px 6px; }
.hero-banner .hero-sub { font-family: 'Fraunces', serif; font-style: italic; font-size: 1.2rem; color: #6b5c46; margin-bottom: 18px; }
.hero-banner .hero-strip {
  display: inline-block; background: #fdf1de; border: 1px solid #f0d9a8; color: #b5591a;
  font-family: 'Fraunces', serif; font-weight: 600; font-size: 1.5rem; padding: 14px 34px;
  border-radius: 999px; letter-spacing: .01em;
}
.hero-welcome { max-width: 780px; margin: 22px auto 0; text-align: justify; text-justify: inter-word; font-size: 1.02rem; line-height: 1.75; color: #4a4232; }
.hero-welcome .eyebrow { text-align: center; font-family: 'JetBrains Mono', monospace; letter-spacing: .25em; font-size: 1.05rem; font-weight: 600; color: #a89a7a; text-transform: uppercase; margin-bottom: 16px; }
.hero-welcome .tagline { font-family: 'Fraunces', serif; font-weight: 700; font-size: 1.4rem; color: var(--coral); margin-top: 22px; }

.site-footer {
  text-align: center; margin-top: 40px; padding: 18px 0;
  border-top: 1px solid #ecdfc4; color: #8a7f68; font-size: .82rem; line-height: 1.7;
}
"""

with gr.Blocks(title="ARIMA-Informed LSTM", theme=CUSTOM_THEME, css=CUSTOM_CSS, head=FONT_HEAD) as demo:
    df_state = gr.State(None)
    best_hp_state = gr.State(None)

    gr.HTML(f'<div id="nav-title"><img src="{_img_to_data_uri("icar_logo.jpg")}" alt="ICAR-IASRI"/></div>')

    with gr.Tabs():
        # ───────────── HOME ─────────────
        with gr.Tab("Home"):
            gr.HTML(
                """
                <div class="hero-banner">
                  <span class="hero-strip">ARIMA-Informed LSTM</span>
                </div>
                """
            )
            gr.HTML(
                """
                <div class="hero-welcome">
                  <div class="eyebrow">— Welcome —</div>
                  <p>The <strong>ARIMA-Informed Long Short-Term Memory (AI-LSTM)</strong> model is a statistically informed forecasting framework inspired by the principles of <strong>Physics-Informed Machine Learning (PIML)</strong>, where domain knowledge is incorporated into the learning process to guide model optimization and improve generalization. In classical PIML, physical laws represented through differential equations are embedded within neural networks as constraints. Analogously, in time-series forecasting, the ARIMA model provides a well-established statistical representation of linear temporal dynamics, including trend, autocorrelation, and persistence structures. The proposed AI-LSTM framework treats the ARIMA model as a source of <strong>statistical prior knowledge</strong>, which is integrated into the LSTM learning process through a composite loss function. While the LSTM component captures complex nonlinear and long-range temporal dependencies, the ARIMA-informed constraint preserves the underlying linear dynamics of the series. By combining statistical knowledge with data-driven learning, the proposed AI-LSTM model follows the philosophy of physics-informed machine learning, enabling more robust, interpretable, and accurate forecasting compared with purely data-driven approaches.</p>
                  
                </div>
                """
            )

        # ───────────── MODEL ─────────────
        with gr.Tab("Model"):
            with gr.Tabs() as model_tabs:
                # -- Data Upload --
                with gr.Tab("1. Data", id=0):
                    gr.Markdown("Upload a CSV or Excel file containing a **time** column and a **study variable** column.")
                    file_in = gr.File(label="Upload Data", file_types=[".csv", ".xlsx", ".xls"])
                    load_msg = gr.Markdown(value="Upload a file above to begin.")
                    with gr.Row():
                        time_col = gr.Dropdown(label="Time Column", choices=[])
                        value_col = gr.Dropdown(label="Study Variable", choices=[])
                    file_in.change(load_file, inputs=file_in,
                                    outputs=[df_state, time_col, value_col, load_msg])
                    gr.Examples(
                        examples=[[os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data.csv")]],
                        inputs=file_in,
                        outputs=[df_state, time_col, value_col, load_msg],
                        fn=load_file,
                        cache_examples=False,
                        label="Or try it with sample data",
                    )
                    next_to_stats_btn = gr.Button("Next → Summary Statistics", variant="secondary")
                    next_to_stats_btn.click(lambda: gr.Tabs(selected=1), outputs=model_tabs)

                # -- Summary stats --
                with gr.Tab("2. Summary Statistics", id=1):
                    select_all_stats = gr.Checkbox(label="Select All", value=False)
                    stats_pick = gr.CheckboxGroup(STAT_CHOICES, value=["Mean", "Median", "Standard Deviation"],
                                                   label="Statistics to display")

                    def _toggle_all_stats(checked):
                        return gr.update(value=STAT_CHOICES if checked else [])

                    select_all_stats.change(_toggle_all_stats, inputs=select_all_stats, outputs=stats_pick)
                    with gr.Row():
                        show_line = gr.Checkbox(label="Show Line Plot", value=True)
                        show_box = gr.Checkbox(label="Show Boxplot", value=True)
                    with gr.Accordion("Plot Size Adjuster", open=False):
                        with gr.Row():
                            plot_width_in = gr.Slider(4, 16, value=8, step=0.5, label="Plot Width")
                            plot_height_in = gr.Slider(2.5, 10, value=3.5, step=0.5, label="Plot Height")
                            font_size_in = gr.Slider(6, 20, value=9, step=1, label="Axis Number Font Size")
                    stats_btn = gr.Button("Compute", variant="primary")
                    stats_out = gr.Markdown(value="Click **Compute** to see statistics here.")
                    line_plot = gr.Plot(label="Time Series")
                    box_plot = gr.Plot(label="Boxplot")
                    stats_btn.click(summary_stats,
                                     inputs=[df_state, time_col, value_col, stats_pick, show_line, show_box,
                                             plot_width_in, plot_height_in, font_size_in],
                                     outputs=[stats_out, line_plot, box_plot])
                    next_to_arima_btn = gr.Button("Next → ARIMA", variant="secondary")
                    next_to_arima_btn.click(lambda: gr.Tabs(selected=2), outputs=model_tabs)

                # -- ARIMA --
                with gr.Tab("3. ARIMA", id=2):
                    train_ratio = gr.Slider(0.5, 0.95, value=0.8, step=0.05, label="Train / Test Split Ratio")
                    use_auto = gr.Checkbox(label="Use auto_arima (auto order selection)", value=True)
                    with gr.Row():
                        p_in = gr.Number(label="p", value=1, precision=0)
                        d_in = gr.Number(label="d", value=1, precision=0)
                        q_in = gr.Number(label="q", value=1, precision=0)
                    arima_btn = gr.Button("Fit ARIMA", variant="primary")
                    arima_out = gr.Markdown(value="Click **Fit ARIMA** to see results here.")
                    arima_result_state = gr.State(None)
                    arima_btn.click(run_arima,
                                     inputs=[df_state, value_col, use_auto, p_in, d_in, q_in, train_ratio],
                                     outputs=[arima_out, arima_result_state])
                    next_to_lstm_btn = gr.Button("Next → LSTM Hyperparameters", variant="secondary")
                    next_to_lstm_btn.click(lambda: gr.Tabs(selected=3), outputs=model_tabs)

                # -- LSTM hyperparameters --
                with gr.Tab("4. LSTM Hyperparameters", id=3):
                    seq_len = gr.Slider(2, 100, value=4, step=1, label="Sequence Length (lag window)")
                    tune_mode = gr.Radio(["Manual", "Grid Search", "Bayesian Search"],
                                          value="Manual", label="Hyperparameter Selection")

                    with gr.Group(visible=True) as manual_group:
                        gr.Markdown("**Manual** — set exact hyperparameter values.")
                        with gr.Row():
                            hidden_in = gr.Number(label="Hidden Size", value=32, precision=0)
                            layers_in = gr.Number(label="Num Layers", value=1, precision=0)
                            lr_in = gr.Number(label="Learning Rate", value=0.003)
                            dropout_in = gr.Number(label="Dropout", value=0.1)

                    with gr.Group(visible=False) as grid_group:
                        gr.Markdown("**Grid Search** — set a min, max, and step for each hyperparameter; "
                                    "every combination in the resulting grid is tried.")
                        with gr.Row():
                            g_hidden_min = gr.Number(label="Hidden Size — Min", value=8, precision=0)
                            g_hidden_max = gr.Number(label="Hidden Size — Max", value=40, precision=0)
                            g_hidden_step = gr.Number(label="Hidden Size — Step", value=8, precision=0)
                        with gr.Row():
                            g_layers_min = gr.Number(label="Num Layers — Min", value=1, precision=0)
                            g_layers_max = gr.Number(label="Num Layers — Max", value=2, precision=0)
                            g_layers_step = gr.Number(label="Num Layers — Step", value=1, precision=0)
                        with gr.Row():
                            g_lr_min = gr.Number(label="Learning Rate — Min", value=0.001)
                            g_lr_max = gr.Number(label="Learning Rate — Max", value=0.005)
                            g_lr_step = gr.Number(label="Learning Rate — Step", value=0.002)
                        with gr.Row():
                            g_dropout_min = gr.Number(label="Dropout — Min", value=0.0)
                            g_dropout_max = gr.Number(label="Dropout — Max", value=0.2)
                            g_dropout_step = gr.Number(label="Dropout — Step", value=0.1)

                    with gr.Group(visible=False) as bayes_group:
                        gr.Markdown("**Bayesian Search** — set a min/max range for each hyperparameter. "
                                    "Uses a Gaussian Process surrogate model with an expected-improvement "
                                    "acquisition function (scikit-optimize `gp_minimize`) to choose each "
                                    "next evaluation based on prior results — true Bayesian optimization, "
                                    "not random sampling.")
                        with gr.Row():
                            b_hidden_min = gr.Number(label="Hidden Size — Min", value=8, precision=0)
                            b_hidden_max = gr.Number(label="Hidden Size — Max", value=40, precision=0)
                        with gr.Row():
                            b_layers_min = gr.Number(label="Num Layers — Min", value=1, precision=0)
                            b_layers_max = gr.Number(label="Num Layers — Max", value=2, precision=0)
                        with gr.Row():
                            b_lr_min = gr.Number(label="Learning Rate — Min", value=0.0005)
                            b_lr_max = gr.Number(label="Learning Rate — Max", value=0.006)
                        with gr.Row():
                            b_dropout_min = gr.Number(label="Dropout — Min", value=0.0)
                            b_dropout_max = gr.Number(label="Dropout — Max", value=0.3)
                        n_calls_in = gr.Slider(5, 40, value=15, step=1, label="Search Iterations")

                    def _toggle_tune_panels(mode):
                        return (gr.update(visible=mode == "Manual"),
                                gr.update(visible=mode == "Grid Search"),
                                gr.update(visible=mode == "Bayesian Search"))

                    tune_mode.change(_toggle_tune_panels, inputs=tune_mode,
                                      outputs=[manual_group, grid_group, bayes_group])

                    lstm_btn = gr.Button("Set / Tune Hyperparameters", variant="primary")
                    lstm_out = gr.Markdown(value="Click **Set / Tune Hyperparameters** to begin.")
                    lstm_btn.click(run_lstm_tuning,
                                    inputs=[df_state, value_col, seq_len, train_ratio, tune_mode,
                                            hidden_in, layers_in, lr_in, dropout_in,
                                            g_hidden_min, g_hidden_max, g_hidden_step,
                                            g_layers_min, g_layers_max, g_layers_step,
                                            g_lr_min, g_lr_max, g_lr_step,
                                            g_dropout_min, g_dropout_max, g_dropout_step,
                                            b_hidden_min, b_hidden_max,
                                            b_layers_min, b_layers_max,
                                            b_lr_min, b_lr_max,
                                            b_dropout_min, b_dropout_max,
                                            n_calls_in],
                                    outputs=[lstm_out, best_hp_state])
                    next_to_ginn_btn = gr.Button("Next → ARIMA-Informed LSTM", variant="secondary")
                    next_to_ginn_btn.click(lambda: gr.Tabs(selected=4), outputs=model_tabs)

                # -- ARIMA-Informed LSTM --
                with gr.Tab("5. ARIMA-Informed LSTM", id=4):
                    gr.Markdown("Runs the LSTM for **λ = 0 (standard)** plus each λ you list below, "
                                "using the ARIMA order and LSTM hyperparameters set in the previous steps.")
                    lambdas_in = gr.Textbox(label="λ values (comma-separated)",
                                             value="0.1, 0.3, 0.5, 0.7, 0.9")
                    with gr.Row():
                        epochs_in = gr.Number(label="Max Epochs", value=200, precision=0)
                        patience_in = gr.Number(label="Early-Stopping Patience", value=20, precision=0)
                    run_btn = gr.Button("Run ARIMA-Informed LSTM", variant="primary")
                    run_status = gr.Markdown(value="Click **Run ARIMA-Informed LSTM** to begin.")
                    gr.Markdown("### Configuration & Hyperparameters Used")
                    config_table = gr.Dataframe(label="Configuration", headers=["Parameter", "Value"])
                    results_table = gr.HTML(label="Results (Train & Test Metrics)")
                    results_plot = gr.Plot(label="Predictions")
                    gr.Markdown("### Actual vs. Predicted — by λ")
                    with gr.Accordion("Training Set: Actual vs Predicted", open=False):
                        train_pred_table = gr.HTML()
                    with gr.Accordion("Test Set: Actual vs Predicted", open=True):
                        test_pred_table = gr.HTML()
                    download_file = gr.File(label="Download Excel Results")
                    run_btn.click(run_full_pipeline,
                                  inputs=[df_state, value_col, seq_len, train_ratio, best_hp_state,
                                          lambdas_in, epochs_in, patience_in],
                                  outputs=[run_status, results_table, results_plot, download_file,
                                           train_pred_table, test_pred_table, config_table])

        # ───────────── INSTRUCTIONS ─────────────
        with gr.Tab("Instructions"):
            gr.HTML(
                """
                <div class="instr-section">
                  <h3>How to use this tool</h3>
                  <div class="step-grid">
                    <div class="step-card">
                      <div class="step-num">1</div>
                      <div class="step-title">Data</div>
                      <div class="step-body">Upload a CSV/Excel file, then pick the time column and the study variable.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">2</div>
                      <div class="step-title">Summary Statistics</div>
                      <div class="step-body">Choose which statistics and plots (line, boxplot) to view.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">3</div>
                      <div class="step-title">ARIMA</div>
                      <div class="step-body">Choose auto order selection or set p, d, q manually. Set the train/test split.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">4</div>
                      <div class="step-title">LSTM Hyperparameters</div>
                      <div class="step-body">Choose manual values, or auto-tune with grid search or Bayesian search.</div>
                    </div>
                    <div class="step-card">
                      <div class="step-num">5</div>
                      <div class="step-title">ARIMA-Informed LSTM</div>
                      <div class="step-body">List the λ values to test - trains a standard LSTM (λ=0) plus one model per λ according to the loss function of (1-λ)*LSTM loss + λ*ARIMA guided loss, reporting RMSE / MAE / MAPE, with a downloadable Excel workbook.</div>
                    </div>
                  </div>
                </div>

                <div class="instr-section">
                  <h3>Methodology</h3>
                  <div class="methodology-box">
                    <p>The proposed ARIMA-Informed Long Short-Term Memory (AI-LSTM) model integrates the
                    linear modelling capability of the Autoregressive Integrated Moving Average (ARIMA)
                    model with the nonlinear sequence learning capability of LSTM networks. Traditional
                    ARIMA models effectively capture linear temporal dependencies and autocorrelation
                    structures in time-series data but often fail to represent complex nonlinear
                    relationships. Conversely, LSTM networks are highly effective in learning nonlinear
                    temporal patterns but do not explicitly incorporate the underlying linear dynamics
                    of the series. To address these limitations, the AI-LSTM framework incorporates
                    ARIMA-derived information directly into the LSTM learning process. First, an optimal
                    ARIMA model is fitted to estimate the linear component and generate one-step-ahead
                    forecasts. These ARIMA forecasts are then used as statistical guidance during LSTM
                    training. The model is optimized using a composite loss function that simultaneously
                    minimizes the forecasting error with respect to the observed values and the deviation
                    from the ARIMA forecasts:</p>

                    <div class="formula-block">
                      L<sub>AI&minus;LSTM</sub> &nbsp;=&nbsp; (1&nbsp;&minus;&nbsp;λ)&nbsp;L<sub>TS</sub> &nbsp;+&nbsp; λL<sub>ARIMA</sub>
                    </div>

                    <p>where</p>
                    <div class="formula-block formula-block--small">
                      L<sub>TS</sub> &nbsp;=&nbsp; <span class="frac"><span class="frac-num">1</span><span class="frac-den">N</span></span>
                      &nbsp;<span class="sum">Σ<sub>t=1</sub><sup>N</sup></span>
                      (y<sub>t</sub> &minus; ŷ<sub>L,t</sub>)<sup>2</sup>
                    </div>
                    <p>represents the forecasting loss based on the observed time-series values, and</p>
                    <div class="formula-block formula-block--small">
                      L<sub>ARIMA</sub> &nbsp;=&nbsp; <span class="frac"><span class="frac-num">1</span><span class="frac-den">N</span></span>
                      &nbsp;<span class="sum">Σ<sub>t=1</sub><sup>N</sup></span>
                      (ŷ<sub>A,t</sub> &minus; ŷ<sub>L,t</sub>)<sup>2</sup>
                    </div>
                    <p>represents the ARIMA consistency loss. Here, y<sub>t</sub> denotes the observed
                    value, ŷ<sub>A,t</sub> denotes the ARIMA forecast, and ŷ<sub>L,t</sub> denotes the
                    LSTM forecast. The weighting parameter λ controls the relative contribution of
                    empirical observations and ARIMA-derived information during model optimization. By
                    combining linear statistical structure and nonlinear deep learning representations
                    within a unified framework, the proposed AI-LSTM model provides a statistically
                    informed forecasting approach capable of improving prediction accuracy for complex
                    time-series data.</p>
                  </div>
                </div>

                
                """
            )

        # ───────────── DEVELOPERS ─────────────
        with gr.Tab("Developers"):
            gr.Markdown("## Developers")
            gr.HTML(build_developers_html())

    footer_html = gr.HTML()
    demo.load(
        lambda: (f"<div class='site-footer'>"
                 f"Copyright © 2026 ICAR – Indian Agricultural Statistics Research Institute, "
                 f"New Delhi - 110012. All Rights Reserved.<br>"
                 f"👥 Total Visitors: {get_and_increment_visitor_count()}</div>"),
        outputs=footer_html,
    )

if __name__ == "__main__":
    import os
    demo.queue().launch(
    server_name="0.0.0.0",
    server_port=int(os.environ.get("PORT", 7860)),
    share=True,
    show_api=False
)